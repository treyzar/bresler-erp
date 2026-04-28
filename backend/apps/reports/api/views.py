import io

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.reports.registry import discover, get_all_reports, get_report


class ReportListView(APIView):
    """GET /api/reports/ — list all available reports."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        discover()
        reports = get_all_reports()
        return Response([r.get_meta() for r in reports])


class ReportDetailView(APIView):
    """GET /api/reports/{name}/ — execute a report with filters.

    Add ?format=xlsx to download as Excel.
    """

    permission_classes = [IsAuthenticated]

    def get(self, request, name):
        discover()
        report = get_report(name)
        if report is None:
            return Response({"detail": "Отчёт не найден"}, status=404)

        # Extract filters from query params
        filters = {}
        for f in report.filters:
            value = request.query_params.get(f.name)
            if value:
                filters[f.name] = value

        data = report.get_data(filters)
        total = len(data)

        # Excel export — all data, no pagination
        if request.query_params.get("export") == "xlsx":
            return self._export_xlsx(report, data)

        # Server-side pagination
        page = int(request.query_params.get("page", 1))
        page_size = int(request.query_params.get("page_size", 50))
        start = (page - 1) * page_size
        end = start + page_size
        page_data = data[start:end]

        meta = report.get_meta()
        return Response(
            {
                "meta": meta,
                "data": page_data,
                "count": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size if page_size else 1,
            }
        )

    @staticmethod
    def _export_xlsx(report, data):
        wb = Workbook()
        ws = wb.active
        ws.title = report.title[:31]  # Excel sheet name limit

        # Header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, col_def in enumerate(report.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.label)
            cell.font = header_font
            cell.fill = header_fill

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_def in enumerate(report.columns, 1):
                value = row_data.get(col_def.name, "")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col_idx, col_def in enumerate(report.columns, 1):
            max_len = len(col_def.label)
            for row_idx in range(2, len(data) + 2):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{report.name}.xlsx"'
        return response
