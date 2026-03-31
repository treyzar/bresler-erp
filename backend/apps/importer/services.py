"""
Import service — parse files, auto-map columns, validate, and apply data.
"""

import csv
import io
import logging

from apps.importer.models import ImportSession

logger = logging.getLogger("importer")

# Target model → {field_name: verbose_name} for auto-mapping
MODEL_FIELDS = {
    "orgunit": {
        "name": "Наименование",
        "full_name": "Полное наименование",
        "unit_type": "Тип",
        "business_role": "Роль",
        "inn": "ИНН",
        "kpp": "КПП",
        "ogrn": "ОГРН",
        "address": "Адрес",
        "external_code": "Внешний код",
        "is_active": "Активен",
    },
    "contact": {
        "full_name": "ФИО",
        "position": "Должность",
        "email": "Email",
        "phone": "Телефон",
        "address": "Адрес",
        "company": "Компания",
    },
    "equipment": {
        "name": "Наименование",
    },
    "typeofwork": {
        "name": "Наименование",
    },
    "facility": {
        "name": "Наименование",
        "address": "Адрес",
    },
}


def parse_file(session: ImportSession) -> list[str]:
    """
    Parse uploaded CSV/XLSX file and extract column headers.
    Returns list of column names.
    """
    filename = session.original_filename.lower()
    session.file.seek(0)

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        columns = _parse_xlsx(session.file)
    else:
        columns = _parse_csv(session.file)

    session.columns = columns
    session.status = ImportSession.Status.MAPPING
    session.save(update_fields=["columns", "status"])
    return columns


def _parse_csv(file) -> list[str]:
    """Parse CSV and return header columns."""
    content = file.read()
    if isinstance(content, bytes):
        # Try UTF-8 with BOM, then cp1251 (common for Russian Excel)
        for encoding in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode("utf-8", errors="replace")
    else:
        text = content

    reader = csv.reader(io.StringIO(text))
    header = next(reader, [])
    return [col.strip() for col in header if col.strip()]


def _parse_xlsx(file) -> list[str]:
    """Parse XLSX and return header columns."""
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True)
    ws = wb.active
    header = []
    for cell in next(ws.iter_rows(max_row=1, values_only=True), []):
        if cell is not None:
            header.append(str(cell).strip())
        else:
            break
    wb.close()
    return header


def auto_map_columns(session: ImportSession) -> dict[str, str]:
    """
    Auto-map file columns to model fields by matching verbose names.
    Returns {file_column: model_field}.
    """
    fields = MODEL_FIELDS.get(session.target_model, {})
    # Build reverse map: lowercase verbose_name → field_name
    reverse = {}
    for field_name, verbose in fields.items():
        reverse[verbose.lower()] = field_name
        # Also map field_name directly
        reverse[field_name.lower()] = field_name

    mapping = {}
    for col in session.columns:
        col_lower = col.strip().lower()
        if col_lower in reverse:
            mapping[col] = reverse[col_lower]

    session.column_mapping = mapping
    session.save(update_fields=["column_mapping"])
    return mapping


def get_available_fields(target_model: str) -> list[dict]:
    """Return available fields for a target model."""
    fields = MODEL_FIELDS.get(target_model, {})
    return [{"name": k, "label": v} for k, v in fields.items()]


def validate_data(session: ImportSession) -> dict:
    """
    Validate all rows using the column mapping. Dry-run — no DB writes.
    Returns {valid_count, error_count, errors: [{row, field, message}], preview: [...]}
    """
    session.status = ImportSession.Status.VALIDATING
    session.save(update_fields=["status"])

    rows = _read_all_rows(session)
    mapping = session.column_mapping
    model_class = _get_model_class(session.target_model)
    required_fields = _get_required_fields(session.target_model)

    valid_rows = []
    errors = []

    for i, row in enumerate(rows, start=2):  # Row 2+ (1 is header)
        mapped_row = {}
        for file_col, model_field in mapping.items():
            if file_col in row:
                mapped_row[model_field] = row[file_col]

        row_errors = _validate_row(mapped_row, required_fields, model_class, i)
        if row_errors:
            errors.extend(row_errors)
        else:
            valid_rows.append(mapped_row)

    total = len(rows)
    session.total_rows = total
    session.error_count = len(set(e["row"] for e in errors))
    session.error_details = errors[:100]  # Limit stored errors
    session.save(update_fields=["total_rows", "error_count", "error_details"])

    # Preview: first 10 valid rows
    preview = valid_rows[:10]

    return {
        "total_rows": total,
        "valid_count": total - session.error_count,
        "error_count": session.error_count,
        "errors": errors[:50],
        "preview": preview,
    }


def apply_import(session: ImportSession) -> dict:
    """
    Apply the import — create records in the database.
    """
    session.status = ImportSession.Status.PROCESSING
    session.save(update_fields=["status"])

    rows = _read_all_rows(session)
    mapping = session.column_mapping
    model_class = _get_model_class(session.target_model)
    required_fields = _get_required_fields(session.target_model)

    success = 0
    errors = []

    for i, row in enumerate(rows, start=2):
        mapped_row = {}
        for file_col, model_field in mapping.items():
            if file_col in row:
                value = row[file_col]
                if value == "" or value is None:
                    continue
                mapped_row[model_field] = value

        row_errors = _validate_row(mapped_row, required_fields, model_class, i)
        if row_errors:
            errors.extend(row_errors)
            continue

        try:
            # Clean boolean fields
            _coerce_types(mapped_row, session.target_model)
            model_class.objects.create(**mapped_row)
            success += 1
        except Exception as exc:
            errors.append({"row": i, "field": "", "message": str(exc)[:200]})

    session.status = ImportSession.Status.COMPLETE
    session.success_count = success
    session.error_count = len(set(e["row"] for e in errors))
    session.error_details = errors[:100]
    session.save(update_fields=["status", "success_count", "error_count", "error_details"])

    # Trigger notification
    from apps.core.events import trigger_event
    trigger_event("import.completed", instance=session, user=session.user)

    logger.info("Import complete: %d success, %d errors", success, session.error_count)
    return {
        "success_count": success,
        "error_count": session.error_count,
        "errors": errors[:50],
    }


def _read_all_rows(session: ImportSession) -> list[dict]:
    """Read all data rows from the uploaded file as list of dicts."""
    filename = session.original_filename.lower()
    session.file.seek(0)

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        return _read_xlsx_rows(session.file)
    return _read_csv_rows(session.file)


def _read_csv_rows(file) -> list[dict]:
    content = file.read()
    if isinstance(content, bytes):
        for encoding in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            text = content.decode("utf-8", errors="replace")
    else:
        text = content

    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx_rows(file) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(file, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c else f"col_{i}" for i, c in enumerate(next(rows_iter, []))]

    result = []
    for row_values in rows_iter:
        row_dict = {}
        for col_name, value in zip(header, row_values):
            row_dict[col_name] = str(value).strip() if value is not None else ""
        result.append(row_dict)

    wb.close()
    return result


def _get_model_class(target_model: str):
    """Get Django model class from target_model string."""
    from apps.directory.models import Contact, Equipment, Facility, OrgUnit, TypeOfWork

    models_map = {
        "orgunit": OrgUnit,
        "contact": Contact,
        "equipment": Equipment,
        "typeofwork": TypeOfWork,
        "facility": Facility,
    }
    return models_map[target_model]


def _get_required_fields(target_model: str) -> set[str]:
    """Get required field names for a target model."""
    required = {
        "orgunit": {"name"},
        "contact": {"full_name"},
        "equipment": {"name"},
        "typeofwork": {"name"},
        "facility": {"name"},
    }
    return required.get(target_model, set())


def _validate_row(mapped_row: dict, required_fields: set, model_class, row_num: int) -> list[dict]:
    """Validate a single row. Returns list of error dicts."""
    errors = []
    for field in required_fields:
        value = mapped_row.get(field, "")
        if not value or (isinstance(value, str) and not value.strip()):
            errors.append({
                "row": row_num,
                "field": field,
                "message": f"Обязательное поле '{field}' пустое",
            })
    return errors


def _coerce_types(mapped_row: dict, target_model: str):
    """Convert string values to appropriate Python types."""
    bool_fields = {"is_active"}
    for field in bool_fields:
        if field in mapped_row:
            val = str(mapped_row[field]).lower().strip()
            mapped_row[field] = val in ("true", "1", "да", "yes", "активен")
