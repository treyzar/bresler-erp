"""
ExportMixin — adds CSV/Excel export to any DRF ViewSet.

Inspired by InvenTree's DataExportMixin. Adds an `export` action that
exports the current queryset (with applied filters) to CSV or XLSX.

Usage:
    class OrderViewSet(ExportMixin, ModelViewSet):
        export_fields = {
            "order_number": "Номер заказа",
            "status": "Статус",
            "customer_org_unit__name": "Заказчик",
            "ship_date": "Дата отгрузки",
        }
        export_filename = "orders"
"""

import csv
import io
from datetime import date, datetime

from django.http import HttpResponse
from rest_framework.decorators import action


class ExportMixin:
    """
    Mixin that adds GET /export/?format=xlsx|csv to a ViewSet.

    Attributes:
        export_fields: dict of {field_lookup: "Column Header"}.
                       Supports nested lookups like "customer_org_unit__name".
                       If None, uses all fields from the list serializer.
        export_filename: Base filename (without extension). Defaults to model name.
    """

    export_fields: dict[str, str] | None = None
    export_filename: str | None = None

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Export filtered queryset to CSV or XLSX."""
        fmt = request.query_params.get("export_format", "xlsx").lower()
        queryset = self.filter_queryset(self.get_queryset())

        fields = self._get_export_fields()
        rows = self._extract_rows(queryset, fields)
        filename = self.export_filename or self._get_model_name()

        if fmt == "csv":
            return self._render_csv(filename, fields, rows)
        return self._render_xlsx(filename, fields, rows)

    def _get_export_fields(self) -> dict[str, str]:
        """Get field mapping {lookup: header}."""
        if self.export_fields:
            return self.export_fields

        # Fallback: extract from model's fields
        model = self.get_queryset().model
        fields = {}
        for f in model._meta.get_fields():
            if hasattr(f, "verbose_name") and not f.many_to_many and not f.one_to_many:
                fields[f.name] = str(f.verbose_name).capitalize()
        return fields

    def _get_model_name(self) -> str:
        return self.get_queryset().model._meta.model_name

    def _extract_rows(self, queryset, fields: dict) -> list[list]:
        """Extract data rows from queryset using field lookups."""
        # Use .values() for efficiency with nested lookups
        lookups = list(fields.keys())

        # Separate simple fields from lookups with __
        simple_fields = [f for f in lookups if "__" not in f]
        nested_fields = [f for f in lookups if "__" in f]
        all_fields = simple_fields + nested_fields

        try:
            qs_values = queryset.values(*all_fields)
        except Exception:
            # If values() fails (e.g., some field doesn't exist), fall back to iteration
            return self._extract_rows_fallback(queryset, lookups)

        rows = []
        for obj in qs_values:
            row = []
            for lookup in lookups:
                value = obj.get(lookup)
                row.append(self._format_value(value))
            rows.append(row)
        return rows

    def _extract_rows_fallback(self, queryset, lookups: list) -> list[list]:
        """Fallback: extract rows by attribute access on model instances."""
        rows = []
        for obj in queryset.iterator():
            row = []
            for lookup in lookups:
                value = self._resolve_lookup(obj, lookup)
                row.append(self._format_value(value))
            rows.append(row)
        return rows

    @staticmethod
    def _resolve_lookup(obj, lookup: str):
        """Resolve a dotted/double-underscore lookup on a model instance."""
        parts = lookup.split("__")
        current = obj
        for part in parts:
            if current is None:
                return None
            current = getattr(current, part, None)
        return current

    @staticmethod
    def _format_value(value) -> str:
        """Format a value for export."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Да" if value else "Нет"
        if isinstance(value, datetime):
            return value.strftime("%d.%m.%Y %H:%M")
        if isinstance(value, date):
            return value.strftime("%d.%m.%Y")
        return str(value)

    def _render_csv(self, filename: str, fields: dict, rows: list) -> HttpResponse:
        """Render data as CSV file."""
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'

        # BOM for Excel compatibility
        response.write("\ufeff")

        writer = csv.writer(response)
        writer.writerow(fields.values())  # Header
        writer.writerows(rows)

        return response

    def _render_xlsx(self, filename: str, fields: dict, rows: list) -> HttpResponse:
        """Render data as XLSX file."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self._render_csv(filename, fields, rows)

        wb = Workbook()
        ws = wb.active
        ws.title = filename[:31]  # Excel sheet name max 31 chars

        # Header row
        headers = list(fields.values())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row in rows:
            ws.append(row)

        # Auto-width columns
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header)
            for row in rows[:100]:  # Sample first 100 rows
                cell_val = row[col_idx - 1] if col_idx - 1 < len(row) else ""
                max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 50)

        # Write to response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
        return response
